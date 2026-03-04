import streamlit as st
import openpyxl
import pandas as pd
from pathlib import Path

EXCEL_PATH = Path(__file__).parent / "Course comparison example for transfer apps.xlsx"


@st.cache_data
def load_articulation_data():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Year tabs: {year: {(school_lower, course_lower): berkeley_equiv}}
    year_tabs = {}
    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            continue
        ws = wb[sheet_name]
        lookup = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            school, course, equiv = row[0], row[1], row[2]
            if school and course and equiv:
                key = (str(school).strip().lower(), str(course).strip().lower())
                lookup[key] = str(equiv).strip()
        year_tabs[year] = lookup

    # Berkeley_REQ: load requirements and series membership
    ws_req = wb["Berkeley_REQ"]
    requirements = []
    req_series = {}    # {berkeley_course: "Yes, chemistry series" | "No"}
    series_groups = {} # {series_name: [berkeley_course, ...]}
    for row in ws_req.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        course = str(row[0]).strip()
        series_val = str(row[1]).strip() if row[1] else "No"
        requirements.append(course)
        req_series[course] = series_val
        if series_val.lower().startswith("yes,"):
            # "Yes, chemistry series" → "chemistry series"
            series_name = series_val[4:].strip()
            series_groups.setdefault(series_name, []).append(course)

    # Berkeley_StronglyRec
    ws_rec = wb["Berkeley_StronglyRec"]
    strongly_rec = [
        str(row[0]).strip()
        for row in ws_rec.iter_rows(min_row=2, values_only=True)
        if row[0]
    ]

    return year_tabs, requirements, req_series, series_groups, strongly_rec


def lookup_equiv(year, school, course, year_tabs):
    if year not in year_tabs:
        return "NOT ARTICULATED"
    key = (str(school).strip().lower(), str(course).strip().lower())
    return year_tabs[year].get(key, "NOT ARTICULATED")


def get_series_label(berkeley_equiv, req_series):
    """Return a human-readable series label for a Berkeley equivalent."""
    if not berkeley_equiv or berkeley_equiv == "NOT ARTICULATED":
        return ""
    val = req_series.get(berkeley_equiv)
    if not val:
        return "N/A"  # course not in Berkeley_REQ (e.g. strongly recommended only)
    if val.lower().startswith("yes,"):
        return val[4:].strip()  # e.g. "chemistry series"
    return val  # "No"


def run_checks(student_df, year_tabs, requirements, req_series, series_groups, strongly_rec):
    df = student_df.copy()

    df["Berkeley Equivalent"] = df.apply(
        lambda r: lookup_equiv(r["Year"], r["School"], r["Course"], year_tabs), axis=1
    )
    df["Part of Series?"] = df["Berkeley Equivalent"].apply(
        lambda e: get_series_label(e, req_series)
    )

    equivs = df["Berkeley Equivalent"].tolist()
    schools = df["School"].tolist()

    # Manual reviews: rows that didn't articulate
    manual_reviews = sum(1 for e in equivs if e == "NOT ARTICULATED")

    # Missing requirements
    missing_reqs = [req for req in requirements if req not in equivs]

    # Series checks: for each series defined in Berkeley_REQ, find which schools
    # the student completed courses in that series at, and flag if mixed.
    series_results = {}
    for series_name, series_courses in series_groups.items():
        matching_schools = {
            str(school).strip()
            for equiv, school in zip(equivs, schools)
            if equiv in series_courses
        }
        if not matching_schools:
            series_results[series_name] = "Courses Not Found"
        elif len(matching_schools) > 1:
            series_results[series_name] = "❌ MIXED SCHOOLS"
        else:
            series_results[series_name] = "✅ OK"

    # Strongly recommended courses the student has covered
    strongly_rec_taken = [c for c in strongly_rec if c in equivs]

    return df, manual_reviews, missing_reqs, series_results, strongly_rec_taken


def parse_paste(text):
    """Parse tab-separated pasted data (Year, School, Course). Header row is skipped."""
    rows = []
    for line in text.strip().splitlines():
        parts = line.split("\t")
        if len(parts) < 3:
            continue
        try:
            year = int(str(parts[0]).strip())
        except ValueError:
            continue  # skip header or malformed rows
        rows.append({
            "Year": year,
            "School": str(parts[1]).strip(),
            "Course": str(parts[2]).strip(),
        })
    return pd.DataFrame(rows) if rows else None


def style_rows(row):
    """Red background + white text for NOT ARTICULATED rows (WCAG AA: ~5:1 contrast)."""
    if row["Berkeley Equivalent"] == "NOT ARTICULATED":
        return ["background-color: #b91c1c; color: #ffffff"] * len(row)
    return [""] * len(row)


# ── UI ────────────────────────────────────────────────────────────────────────

st.title("UC Berkeley Transfer Course Checker")

try:
    year_tabs, requirements, req_series, series_groups, strongly_rec = load_articulation_data()
except FileNotFoundError:
    st.error(f"Excel file not found at: {EXCEL_PATH}")
    st.stop()

st.markdown(
    "Paste your student course data below (columns A–C from StudentData: **Year, School, Course**). "
    "Copy directly from Excel — tab-separated, with or without a header row."
)

pasted = st.text_area(
    "Student data",
    height=220,
    placeholder="2025\tContra Costa\tMath 110\n2025\tContra Costa\tChem 203\n2024\tDiablo Valley\tBiology 1a",
)

if pasted.strip():
    student_df = parse_paste(pasted)
    if student_df is None or student_df.empty:
        st.error(
            "Could not parse the pasted data. Make sure it has Year, School, and Course "
            "columns separated by tabs (copy straight from Excel)."
        )
        st.stop()

    result_df, manual_reviews, missing_reqs, series_results, strongly_rec_taken = run_checks(
        student_df, year_tabs, requirements, req_series, series_groups, strongly_rec
    )

    st.divider()
    st.subheader("Course Lookup")
    st.dataframe(
        result_df.style.apply(style_rows, axis=1),
        use_container_width=True,
        hide_index=True,
    )

    st.divider()

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Manual Reviews Needed")
        if manual_reviews == 0:
            st.success("0 — no manual reviews needed")
        else:
            st.error(f"{manual_reviews} course(s) did not articulate and need manual review")

        st.subheader("Missing Requirements")
        if not missing_reqs:
            st.success("All requirements met!")
        else:
            for req in missing_reqs:
                st.error(req)

        st.subheader("Strongly Recommended Courses Taken / Planned")
        if strongly_rec_taken:
            for course in strongly_rec_taken:
                st.success(course)
        else:
            st.info("None taken or planned")

    with col2:
        st.subheader("Course Series Checks")
        if not series_results:
            st.info("No series defined in Berkeley requirements.")
        for series_name, result in series_results.items():
            label = series_name.title()
            if "OK" in result:
                st.success(f"**{label}**: {result}")
            elif "MIXED" in result:
                st.warning(f"**{label}**: {result}")
            else:
                st.info(f"**{label}**: {result}")
